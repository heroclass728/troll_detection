import os
import openpyxl
import pandas as pd
import datetime

from settings import EXCEL_FILE, SHEET_NAME


class ImportData:

    def __init__(self):
        self.excel_path = EXCEL_FILE
        self.__initiate_excel_file()

    def __initiate_excel_file(self):
        if not os.path.exists(self.excel_path):
            wb = openpyxl.Workbook()
            wb.save(self.excel_path)
            writer = pd.ExcelWriter(self.excel_path, engine='openpyxl', mode='w')
            df = pd.DataFrame(columns=["Date", "Time", "Barcode", "Attr", "Type", "Amount"])
            df.to_excel(writer, sheet_name=SHEET_NAME, index=False)
            writer.save()
            writer.close()

    def import_parsing_data(self, barcode, types, amounts):
        current_dt = [datetime.datetime.now().strftime("%Y-%m-%d")] * len(types)
        current_time = [datetime.datetime.now().strftime("%H-%M-%S")] * len(types)
        barcodes = [barcode] * len(types)
        attrs = [""] * len(types)

        df = pd.DataFrame(list(zip(current_dt, current_time, barcodes, attrs, types, amounts)), columns=["Date", "Time",
                                                                                                         "Barcode",
                                                                                                         "Attr", "Type",
                                                                                                         "Amount"])
        append_df_to_excel(self.excel_path, df, sheet_name=SHEET_NAME, index=False, header=None)

        return


def append_df_to_excel(filename, df, sheet_name, start_row=None, truncate_sheet=False, **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      start_row : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = openpyxl.load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if start_row is None and sheet_name in writer.book.sheetnames:
            start_row = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if start_row is None:
        start_row = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=start_row, **to_excel_kwargs)

    # save the workbook
    writer.save()


if __name__ == '__main__':
    data_importer = ImportData()
    data_importer.import_parsing_data(barcode="", types=["water", "coca cola", "pepsi"], amounts=[2, 2, 2])
